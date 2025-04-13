import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException
import openpyxl
import os
import shutil
import threading
import time
import random
from datetime import datetime
from typing import List, Dict, Optional, Set


class JobinjaScraper:
    """Handles the web scraping functionality for Jobinja website"""
    
    def __init__(self, driver_path: str):
        self.driver_path = driver_path
        self.driver = None
        
    def setup_driver(self) -> Optional[webdriver.Chrome]:
        """Initialize and configure Chrome WebDriver"""
        try:
            options = Options()
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--blink-settings=imagesEnabled=false")
            options.add_argument("--disable-images")
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-logging")
            options.add_argument("--log-level=3")
            options.add_argument("--output=/dev/null")
            options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            
            if not os.path.exists(self.driver_path):
                raise FileNotFoundError(f"Chromedriver not found at: {self.driver_path}")
            
            service = Service(executable_path=self.driver_path)
            driver = webdriver.Chrome(service=service, options=options)
            driver.set_page_load_timeout(15)
            self.driver = driver
            return driver
                
        except WebDriverException as e:
            raise WebDriverException(f"Browser setup error (check VPN/internet): {str(e)}")
        except Exception as e:
            raise Exception(f"Browser setup error: {str(e)}")

    def extract_job_data(self, url: str) -> Dict[str, str]:
        """Extract job data from a given URL"""
        if not self.driver:
            self.driver = self.setup_driver()
            if not self.driver:
                raise RuntimeError("Driver not initialized. Please ensure setup_driver() is called successfully.")
        
        try:
            self.driver.set_page_load_timeout(30)
            self.driver.get(url)
            time.sleep(random.uniform(2, 5))
            
            data = {}
            
            def quick_extract(xpath: str) -> str:
                try:
                    if self.driver is None:
                        raise RuntimeError("Driver initialization failed.")
                    element = self.driver.find_element(By.XPATH, xpath)
                    return str(element.text).strip() if element.text else "N/A"
                except:
                    return "N/A"
            
            data["Job Title"] = quick_extract("//h1")
            data["Category"] = quick_extract('//h4[text()="دسته‌بندی شغلی"]/following-sibling::div/span')
            data["Location"] = quick_extract('//h4[text()="موقعیت مکانی"]/following-sibling::div/span')
            data["Cooperation Type"] = quick_extract('//h4[text()="نوع همکاری"]/following-sibling::div/span')
            data["Work Experience"] = quick_extract('//h4[text()="حداقل سابقه کار"]/following-sibling::div/span')
            data["Salary"] = quick_extract('//h4[text()="حقوق"]/following-sibling::div/span')
            data["Languages"] = quick_extract('//h4[text()="زبان‌های مورد نیاز"]/following-sibling::div/span')
            
            try:
                skills = [str(skill.text).strip() for skill in self.driver.find_elements(
                    By.XPATH, '//h4[text()="مهارت‌های مورد نیاز"]/following-sibling::div/span')]
                data["Skills"] = ", ".join(skills) if skills else "N/A"
            except:
                data["Skills"] = "N/A"
                
            data["Gender"] = quick_extract('//h4[text()="جنسیت"]/following-sibling::div/span')
            data["Military Status"] = quick_extract('//h4[text()="وضعیت نظام وظیفه"]/following-sibling::div/span')
            data["Education Level"] = quick_extract('//h4[text()="حداقل مدرک تحصیلی"]/following-sibling::div/span')

            try:
                desc = self.driver.find_element(By.CSS_SELECTOR, ".o-box__text.s-jobDesc")
                data["Job Description"] = str(desc.text).strip() if desc.text else "N/A"
            except:
                data["Job Description"] = "N/A"

            try:
                company = self.driver.find_element(By.CSS_SELECTOR, ".o-box__text:not(.s-jobDesc)")
                data["Company Introduction"] = str(company.text).strip() if company.text else "N/A"
            except:
                data["Company Introduction"] = "N/A"

            data["URL"] = url

            return data

        except WebDriverException as e:
            raise WebDriverException(f"Connection error for {url} (check VPN): {str(e)}")
        except Exception as e:
            raise Exception(f"Error extracting data from {url}: {str(e)}")
    
    def close(self):
        """Close the WebDriver"""
        if self.driver:
            self.driver.quit()
            self.driver = None


class ExcelHandler:
    """Handles all Excel file operations"""
    
    def __init__(self):
        self.headers = [
            "Job Title", "Category", "Location", "Cooperation Type",
            "Work Experience", "Salary", "Languages", "Skills",
            "Gender", "Military Status", "Education Level",
            "Job Description", "Company Introduction", "URL"
        ]
    
    def get_existing_links(self, file_path: str) -> Set[str]:
        """Get all links from an existing Excel file"""
        existing_links = set()
        
        if not os.path.exists(file_path):
            return existing_links
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            if ws is None:
                return existing_links
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[-1]:
                    link = str(row[-1]).strip()
                    existing_links.add(link)
                    
        except Exception as e:
            raise Exception(f"Error reading existing links: {str(e)}")
        
        return existing_links
    
    def create_new_output_file(self, file_path: str) -> bool:
        """Create a new output file with headers"""
        try:
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise Exception("Failed to create worksheet")
            
            ws.append(self.headers)
            wb.save(file_path)
            return True
            
        except Exception as e:
            raise Exception(f"Error creating new output file: {str(e)}")
    
    def read_input_links(self, file_path: str) -> List[str]:
        """Read job links from input Excel file"""
        links = []
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Input file not found: {file_path}")
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        if ws is None:
            raise Exception("Invalid input worksheet")
        
        for row in ws.iter_rows(min_row=2):
            if row and row[-1] and row[-1].value:
                link = str(row[-1].value).strip()
                links.append(link)
        
        return links
    
    def append_data(self, file_path: str, data: Dict) -> bool:
        """Append new data to the Excel file (preserving existing order)"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            if ws is None:
                raise Exception("No active worksheet found")
            
            # Create new row
            new_row = [data.get(header, "N/A") for header in self.headers]
            
            # Add new data at the end
            ws.append(new_row)
            
            wb.save(file_path)
            return True
            
        except Exception as e:
            raise Exception(f"Error saving data: {str(e)}")
    
    def copy_existing_data_to_new_file(self, source_path: str, target_path: str) -> bool:
        """Copy all data from existing file to new file"""
        try:
            if not os.path.exists(source_path):
                return False

            # Create new file with headers
            self.create_new_output_file(target_path)

            # Open both files
            source_wb = openpyxl.load_workbook(source_path)
            source_ws = source_wb.active
            if source_ws is None:
                raise Exception("شیت فعال در فایل منبع یافت نشد.")

            target_wb = openpyxl.load_workbook(target_path)
            target_ws = target_wb.active
            if target_ws is None:
                raise Exception("شیت فعال در فایل مقصد یافت نشد.")

            # Copy all rows (except headers) from source to target
            for row in source_ws.iter_rows(min_row=2, values_only=True):
                target_ws.append(row)

            # Save the target file
            target_wb.save(target_path)
            return True

        except Exception as e:
            raise Exception(f"Error copying data to new file: {str(e)}")


class JobinjaExcelUpdaterApp:
    """Main application GUI and processing controller"""
    
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Jobinja Scraper - پردازشگر فایل اکسل")
        self.root.geometry("900x750")
        self.root.resizable(False, False)
        
        # Configuration
        self.input_file = ""
        self.existing_output_file = ""
        self.new_output_file = ""
        self.chrome_driver_path = "C:/Users/ASUS/Desktop/chromedriver-win64/chromedriver.exe"
        self.is_running = False
        self.processed_count = 0
        self.total_count = 0
        self.delay_var = tk.IntVar(value=0)  # Added delay_var attribute
        self.is_paused = False  # متغیر برای مدیریت توقف/ادامه
        self.schedule_mode = tk.StringVar(value="Immediate")  # حالت زمان‌بندی
        self.schedule_interval = tk.IntVar(value=2)  # فاصله زمانی (ساعت)
        self.delay_seconds = tk.IntVar(value=2)  # تأخیر بین درخواست‌ها
        
        # Initialize components
        self.scraper = JobinjaScraper(self.chrome_driver_path)
        self.excel_handler = ExcelHandler()
        
        # Setup GUI
        self.create_widgets()
        self.set_styles()
    
    def set_styles(self):
        """Configure GUI styles"""
        style = ttk.Style()
        style.configure("TFrame", background="#f0f0f0")
        style.configure("TLabel", background="#f0f0f0", font=("Arial", 10))
        style.configure("TButton", font=("Arial", 10), padding=5)
        style.configure("TProgressbar", thickness=20)
        style.configure("Title.TLabel", font=("Arial", 16, "bold"))
        style.configure("Status.TLabel", font=("Arial", 9), foreground="#555555")
        style.configure("Red.TButton", foreground="red")
    
    def create_widgets(self):
        """Create and arrange all GUI components"""
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ttk.Label(main_frame, text="برنامه جمع‌آوری اطلاعات جابینجا", style="Title.TLabel")
        title_label.pack(pady=(0, 20))
        
        # File selection section
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(input_frame, text="فایل اکسل ورودی (لیست لینک‌ها):").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.input_entry = ttk.Entry(input_frame, width=60)
        self.input_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(input_frame, text="انتخاب فایل", command=self.choose_input_file).grid(row=0, column=2, padx=5, pady=5)
        
        ttk.Label(input_frame, text="فایل اکسل موجود (داده‌های قبلی):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.existing_output_entry = ttk.Entry(input_frame, width=60)
        self.existing_output_entry.grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(input_frame, text="انتخاب فایل", command=self.choose_existing_output_file).grid(row=1, column=2, padx=5, pady=5)
        
        ttk.Label(input_frame, text="فایل اکسل جدید (خروجی نهایی):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.new_output_entry = ttk.Entry(input_frame, width=60)
        self.new_output_entry.grid(row=2, column=1, padx=5, pady=5)
        ttk.Button(input_frame, text="انتخاب فایل", command=self.choose_new_output_file).grid(row=2, column=2, padx=5, pady=5)
        
        # Progress display
        self.progress_label = ttk.Label(main_frame, text="آماده شروع...", style="Status.TLabel")
        self.progress_label.pack(pady=(10, 5))
        
        self.progress_bar = ttk.Progressbar(main_frame, orient=tk.HORIZONTAL, length=700, mode='determinate')
        self.progress_bar.pack(pady=5)
        
        self.status_label = ttk.Label(main_frame, text="", style="Status.TLabel")
        self.status_label.pack(pady=5)
        
        # Control buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        self.start_button = ttk.Button(button_frame, text="شروع پردازش", command=self.start_processing)
        self.start_button.pack(side=tk.LEFT, padx=10)
        
        self.stop_button = ttk.Button(button_frame, text="توقف", command=self.stop_processing, 
                                    state=tk.DISABLED, style="Red.TButton")
        self.stop_button.pack(side=tk.LEFT, padx=10)
        
        # Pause/Resume Button
        self.pause_button = ttk.Button(button_frame, text="توقف", command=self.toggle_pause, state=tk.DISABLED)
        self.pause_button.pack(side=tk.LEFT, padx=10)

        # Schedule Section
        schedule_frame = ttk.Frame(main_frame)
        schedule_frame.pack(fill=tk.X, pady=10)
        ttk.Label(schedule_frame, text="حالت زمان‌بندی:").grid(row=0, column=0, padx=5, pady=5)
        ttk.OptionMenu(schedule_frame, self.schedule_mode, "Immediate", "Immediate", "Scheduled").grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(schedule_frame, text="فاصله زمانی (ساعت):").grid(row=1, column=0, padx=5, pady=5)
        ttk.Entry(schedule_frame, textvariable=self.schedule_interval, width=10).grid(row=1, column=1, padx=5, pady=5)

        # Delay Setting
        delay_frame = ttk.Frame(main_frame)
        delay_frame.pack(fill=tk.X, pady=10)
        ttk.Label(delay_frame, text="تأخیر بین درخواست‌ها (ثانیه):").grid(row=0, column=0, padx=5, pady=5)
        ttk.Entry(delay_frame, textvariable=self.delay_seconds, width=10).grid(row=0, column=1, padx=5, pady=5)

        # Log section
        log_frame = ttk.Frame(main_frame)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        ttk.Label(log_frame, text="گزارش عملیات:").pack(anchor=tk.W)
        
        self.log_text = tk.Text(log_frame, height=12, wrap=tk.WORD, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(self.log_text)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.log_text.yview)
    
    def choose_input_file(self):
        """Handle input file selection"""
        file_path = filedialog.askopenfilename(
            title="انتخاب فایل اکسل ورودی", 
            filetypes=[("فایل‌های اکسل", "*.xlsx")]
        )
        if file_path:
            self.input_file = file_path
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, file_path)
            self.log_message(f"فایل ورودی انتخاب شد: {file_path}")
    
    def choose_existing_output_file(self):
        """Handle existing output file selection"""
        file_path = filedialog.askopenfilename(
            title="انتخاب فایل اکسل موجود",
            filetypes=[("فایل‌های اکسل", "*.xlsx")]
        )
        if file_path:
            self.existing_output_file = file_path
            self.existing_output_entry.delete(0, tk.END)
            self.existing_output_entry.insert(0, file_path)
            self.log_message(f"فایل اکسل موجود انتخاب شد: {file_path}")
    
    def choose_new_output_file(self):
        """Handle new output file selection/creation"""
        file_path = filedialog.asksaveasfilename(
            title="ایجاد فایل اکسل خروجی جدید",
            defaultextension=".xlsx",
            filetypes=[("فایل‌های اکسل", "*.xlsx")]
        )
        
        if file_path:
            self.new_output_file = file_path
            self.new_output_entry.delete(0, tk.END)
            self.new_output_entry.insert(0, file_path)
            self.log_message(f"فایل خروجی جدید ایجاد خواهد شد: {file_path}")
    
    def log_message(self, message: str):
        """Add a message to the log with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.config(state=tk.DISABLED)
        self.log_text.see(tk.END)
        self.root.update()
    
    def update_progress(self, current: int, total: int):
        """Update progress bar and labels"""
        self.processed_count = current
        self.total_count = total
        progress_percent = int((current / total) * 100) if total > 0 else 0
        
        self.progress_bar["value"] = progress_percent
        self.progress_label.config(text=f"پیشرفت: {current} از {total} ({progress_percent}%)")
        self.status_label.config(text=f"در حال پردازش آیتم {current} از {total}")
        self.root.update()
    
    def start_processing(self):
        """Start the processing thread"""
        if not self.input_file:
            messagebox.showerror("خطا", "لطفاً فایل ورودی را انتخاب کنید")
            return
        
        if not self.existing_output_file:
            messagebox.showerror("خطا", "لطفاً فایل اکسل موجود را انتخاب کنید")
            return
            
        if not self.new_output_file:
            messagebox.showerror("خطا", "لطفاً فایل خروجی جدید را انتخاب یا ایجاد کنید")
            return
        
        if self.is_running:
            return
        
        self.is_running = True
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.pause_button.config(state=tk.NORMAL)
        self.log_message("شروع پردازش اطلاعات...")
        
        threading.Thread(target=self.run_processing, daemon=True).start()
    
    def stop_processing(self):
        """Request processing to stop after current item"""
        self.is_running = False
        self.log_message("درخواست توقف پس از اتمام آیتم جاری...")
        self.stop_button.config(state=tk.DISABLED)
    
    def toggle_pause(self):
        """Toggle pause/resume state"""
        if self.is_paused:
            self.is_paused = False
            self.pause_button.config(text="توقف")
            self.log_message("✅ ادامه عملیات...")
        else:
            self.is_paused = True
            self.pause_button.config(text="ادامه")
            self.log_message("⏸ عملیات متوقف شد.")

    def run_processing(self):
        """Main processing function (runs in separate thread)"""
        try:
            # Copy existing data to new file first
            try:
                self.log_message("در حال کپی داده‌های موجود به فایل جدید...")
                if not self.excel_handler.copy_existing_data_to_new_file(self.existing_output_file, self.new_output_file):
                    messagebox.showerror("خطا", "کپی داده‌های موجود به فایل جدید ناموفق بود")
                    return
                self.log_message("✅ داده‌های موجود با موفقیت به فایل جدید کپی شدند")
            except Exception as e:
                messagebox.showerror("خطا", f"خطا در کپی داده‌های موجود:\n{str(e)}")
                return

            # Setup browser
            try:
                self.scraper.setup_driver()
            except Exception as e:
                messagebox.showerror("خطا", f"راه‌اندازی مرورگر ناموفق بود:\n{str(e)}")
                return

            # Read input links
            try:
                input_links = self.excel_handler.read_input_links(self.input_file)
            except Exception as e:
                self.log_message(f"⚠️ خطا در خواندن فایل ورودی: {str(e)}")
                return
            
            if not input_links:
                self.log_message("⚠️ هیچ لینکی در فایل ورودی یافت نشد!")
                return
            
            self.update_progress(0, len(input_links))
            
            # Get all existing links from output file for duplicate checking
            existing_links = self.excel_handler.get_existing_links(self.existing_output_file)
            self.log_message(f"بررسی تکراری در میان {len(existing_links)} لینک موجود")
            
            # Process each link
            processed_count = 0
            stop_reason = ""
            
            current_row = 2  # شروع از ردیف دوم برای داده‌های جدید

            while True:
                if self.schedule_mode.get() == "Scheduled":
                    self.log_message(f"⏳ عملیات زمان‌بندی شده شروع خواهد شد (هر {self.schedule_interval.get()} ساعت).")
                    time.sleep(self.schedule_interval.get() * 3600)  # تبدیل ساعت به ثانیه

                for i, link in enumerate(input_links):
                    while self.is_paused:
                        time.sleep(1)  # توقف در حالت Pause

                    if not self.is_running:
                        stop_reason = "پردازش توسط کاربر متوقف شد"
                        break

                    # Check for duplicate link
                    if link in existing_links:
                        self.log_message(f"⛔ لینک تکراری شناسایی شد: {link}")
                        raise Exception(f"لینک تکراری شناسایی شد: {link}")  # توقف پردازش در صورت لینک تکراری

                    # ادامه پردازش لینک
                    self.update_progress(i + 1, len(input_links))
                    self.log_message(f"پردازش لینک {i+1}: {link}")

                    attempt = 0
                    data = None
                    while attempt < 3:  # تلاش تا 3 بار
                        try:
                            attempt += 1
                            self.log_message(f"🔄 تلاش {attempt} برای استخراج داده از لینک: {link}")
                            data = self.scraper.extract_job_data(link)
                            if data:
                                self.log_message(f"✅ داده با موفقیت استخراج شد از لینک: {link}")
                                break
                        except Exception as e:
                            self.log_message(f"⚠️ خطا در تلاش {attempt} برای استخراج داده از لینک {link}: {str(e)}")

                    if not data:  # اگر بعد از 3 تلاش داده‌ای استخراج نشد
                        self.log_message(f"❌ داده‌ای از لینک {link} استخراج نشد. ذخیره مقدار پیش‌فرض در اکسل.")
                        data = {
                            "Job Title": "داده استخراج نشد",
                            "Category": "داده استخراج نشد",
                            "Location": "داده استخراج نشد",
                            "Cooperation Type": "داده استخراج نشد",
                            "Work Experience": "داده استخراج نشد",
                            "Salary": "داده استخراج نشد",
                            "Languages": "داده استخراج نشد",
                            "Skills": "داده استخراج نشد",
                            "Gender": "داده استخراج نشد",
                            "Military Status": "داده استخراج نشد",
                            "Education Level": "داده استخراج نشد",
                            "Job Description": "داده استخراج نشد",
                            "Company Introduction": "داده استخراج نشد",
                            "URL": link,
                        }

                    # ذخیره داده در اکسل
                    try:
                        wb_output = openpyxl.load_workbook(self.new_output_file)
                        ws_output = wb_output.active
                        if ws_output is None:
                            raise Exception("شیت فعال در فایل اکسل یافت نشد.")

                        # اضافه کردن یک ردیف جدید در موقعیت current_row
                        ws_output.insert_rows(current_row)

                        # اضافه کردن داده‌ها به ردیف فعلی
                        for col_idx, value in enumerate([
                            data.get("Job Title", "N/A"),
                            data.get("Category", "N/A"),
                            data.get("Location", "N/A"),
                            data.get("Cooperation Type", "N/A"),
                            data.get("Work Experience", "N/A"),
                            data.get("Salary", "N/A"),
                            data.get("Languages", "N/A"),
                            data.get("Skills", "N/A"),
                            data.get("Gender", "N/A"),
                            data.get("Military Status", "N/A"),
                            data.get("Education Level", "N/A"),
                            data.get("Job Description", "N/A"),
                            data.get("Company Introduction", "N/A"),
                            link
                        ], start=1):
                            ws_output.cell(row=current_row, column=col_idx, value=value)

                        # افزایش شمارنده ردیف برای داده بعدی
                        current_row += 1

                        # ذخیره تغییرات در فایل خروجی
                        wb_output.save(self.new_output_file)
                        self.log_message(f"✅ داده با موفقیت در اکسل ذخیره شد برای لینک: {link}")
                        existing_links.add(link)
                    except Exception as e:
                        self.log_message(f"⚠️ خطا در ذخیره داده در اکسل برای لینک {link}: {str(e)}")

                    # Random delay between requests
                    time.sleep(self.delay_seconds.get())

                # پشتیبان‌گیری
                self.save_backup()

                if self.schedule_mode.get() == "Immediate":
                    break

            # Final status message
            if stop_reason:
                self.log_message(f"⏹️ {stop_reason}")
                messagebox.showinfo("توقف پردازش", stop_reason)
            elif self.is_running:
                self.log_message(f"✅ پردازش با موفقیت заверш شد! {processed_count} رکورد جدید اضافه شد")
                messagebox.showinfo("موفقیت", f"پردازش کامل شد. {processed_count} رکورد جدید به فایل اضافه شد.")
            else:
                self.log_message(f"⏹️ پردازش متوقف شد! {processed_count} رکورد پردازش شد")
                messagebox.showinfo("توقف", f"پردازش متوقف شد. {processed_count} رکورد پردازش شد.")
            
        except Exception as e:
            self.log_message(f"⚠️ خطای جدی: {str(e)}")
            messagebox.showerror("خطا", f"خطای غیرمنتظره:\n{str(e)}")
        finally:
            self.scraper.close()
            self.is_running = False
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
            self.pause_button.config(state=tk.DISABLED)
            self.status_label.config(text="آماده شروع مجدد")

    def save_backup(self):
        """Create a backup of the output file"""
        try:
            if not self.new_output_file:
                return
                
            backup_dir = os.path.join(os.path.dirname(self.new_output_file), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"backup_{timestamp}.xlsx")
            
            shutil.copy2(self.new_output_file, backup_path)
            self.log_message(f"✅ نسخه پشتیبان ایجاد شد: {backup_path}")
            
        except Exception as e:
            self.log_message(f"⚠️ خطا در ایجاد نسخه پشتیبان: {str(e)}")

    def save_status(self, processed_count: int, existing_links: Set[str]):
        """Save current processing status"""
        # This can be expanded to save more detailed status if needed
        self.log_message(f"💾 ذخیره وضعیت فعلی: {processed_count} لینک پردازش شده")


def main():
    root = tk.Tk()
    app = JobinjaExcelUpdaterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()