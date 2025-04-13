import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl
import os
import shutil
import threading
import json
from datetime import datetime
import time
import random

class JobinjaScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Jobinja Scraper v2.1 (Optimized)")
        self.root.geometry("900x650")
        self.root.resizable(False, False)
        
        # Variables
        self.input_file = ""
        self.output_file = ""
        self.chrome_driver_path = r"C:\Users\ASUS\Desktop\chromedriver-win64\chromedriver.exe"
        self.is_running = False
        self.processed_count = 0
        self.total_count = 0
        self.status_file = "jobinja_status.json"
        
        self.create_widgets()
        self.set_styles()
    
    def set_styles(self):
        style = ttk.Style()
        style.configure("TFrame", background="#f0f0f0")
        style.configure("TLabel", background="#f0f0f0", font=("Arial", 10))
        style.configure("TButton", font=("Arial", 10), padding=5)
        style.configure("TProgressbar", thickness=20)
        style.configure("Title.TLabel", font=("Arial", 16, "bold"))
        style.configure("Status.TLabel", font=("Arial", 9), foreground="#555555")
        style.configure("Red.TButton", foreground="red")
        
    def create_widgets(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        title_label = ttk.Label(main_frame, text="Jobinja Job Scraper (Fast Mode)", style="Title.TLabel")
        title_label.pack(pady=(0, 20))
        
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(input_frame, text="Input Excel File:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.input_entry = ttk.Entry(input_frame, width=60)
        self.input_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(input_frame, text="Browse", command=self.choose_input_file).grid(row=0, column=2, padx=5, pady=5)
        
        ttk.Label(input_frame, text="Output Excel File:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.output_entry = ttk.Entry(input_frame, width=60)
        self.output_entry.grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(input_frame, text="Browse", command=self.choose_output_file).grid(row=1, column=2, padx=5, pady=5)
        
        self.resume_var = tk.IntVar(value=1)
        ttk.Checkbutton(
            input_frame, 
            text="Resume from last processed", 
            variable=self.resume_var,
            onvalue=1, 
            offvalue=0
        ).grid(row=2, column=1, sticky=tk.W, pady=5)
        
        self.progress_label = ttk.Label(main_frame, text="Ready to start...", style="Status.TLabel")
        self.progress_label.pack(pady=(20, 5))
        
        self.progress_bar = ttk.Progressbar(main_frame, orient=tk.HORIZONTAL, length=700, mode='determinate')
        self.progress_bar.pack(pady=5)
        
        self.status_label = ttk.Label(main_frame, text="", style="Status.TLabel")
        self.status_label.pack(pady=5)
        
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        self.start_button = ttk.Button(button_frame, text="Start Scraping", command=self.start_scraping)
        self.start_button.pack(side=tk.LEFT, padx=10)
        
        self.stop_button = ttk.Button(button_frame, text="Stop", command=self.stop_scraping, state=tk.DISABLED, style="Red.TButton")
        self.stop_button.pack(side=tk.LEFT, padx=10)
        
        log_frame = ttk.Frame(main_frame)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        ttk.Label(log_frame, text="Operation Log:").pack(anchor=tk.W)
        
        self.log_text = tk.Text(log_frame, height=12, wrap=tk.WORD, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(self.log_text)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.log_text.yview)
    
    def choose_input_file(self):
        file_path = filedialog.askopenfilename(title="Select Input Excel File", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.input_file = file_path
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, file_path)
            self.log_message(f"Input file selected: {file_path}")
    
    def choose_output_file(self):
        response = messagebox.askyesno("Output File", "Does the output Excel file already exist?")
        if response:
            file_path = filedialog.askopenfilename(title="Select Output Excel File", filetypes=[("Excel files", "*.xlsx")])
        else:
            file_path = filedialog.asksaveasfilename(title="Save Output Excel File", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        
        if file_path:
            self.output_file = file_path
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, file_path)
            self.log_message(f"Output file selected: {file_path}")
    
    def log_message(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.config(state=tk.DISABLED)
        self.log_text.see(tk.END)
        self.root.update()
    
    def update_progress(self, current, total):
        self.processed_count = current
        self.total_count = total
        progress_percent = int((current / total) * 100) if total > 0 else 0
        
        self.progress_bar["value"] = progress_percent
        self.progress_label.config(text=f"Processing: {current} of {total} ({progress_percent}%)")
        self.status_label.config(text=f"Processing link {current} of {total}")
        self.root.update()
    
    def start_scraping(self):
        if not self.input_file or not self.output_file:
            messagebox.showerror("Error", "Please select both input and output files.")
            return
        
        if self.is_running:
            return
        
        self.is_running = True
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.log_message("Starting data extraction process...")
        
        threading.Thread(target=self.run_scraping, daemon=True).start()
    
    def stop_scraping(self):
        self.is_running = False
        self.log_message("Stop request received...")
        self.stop_button.config(state=tk.DISABLED)
    
    def save_status(self, last_index):
        try:
            status_data = {
                'last_index': last_index,
                'output_file': self.output_file,
                'input_file': self.input_file,
                'timestamp': datetime.now().isoformat()
            }
            with open(self.status_file, "w", encoding='utf-8') as f:
                json.dump(status_data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            self.log_message(f"⚠️ Error saving status: {str(e)}")
    
    def load_status(self):
        try:
            if os.path.exists(self.status_file):
                with open(self.status_file, "r", encoding='utf-8') as f:
                    status = json.load(f)
                    if (status['input_file'] == self.input_file and 
                        os.path.exists(status['output_file'])):
                        return status['last_index']
            return 0
        except Exception as e:
            self.log_message(f"⚠️ Error loading status: {str(e)}")
            return 0
    
    def setup_driver(self):
        try:
            options = Options()
            # options.add_argument("--headless")
            
            # Lightweight settings for maximum speed
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--blink-settings=imagesEnabled=false")
            options.add_argument("--disable-images")
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-logging")
            options.add_argument("--log-level=3")
            options.add_argument("--output=/dev/null")
            
            # Lightweight mobile user agent
            options.add_argument("user-agent=Mozilla/5.0 (Linux; Android 10) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.120 Mobile Safari/537.36")
            
            if not os.path.exists(self.chrome_driver_path):
                self.log_message(f"⚠️ ChromeDriver not found at {self.chrome_driver_path}")
                return None
            
            service = Service(executable_path=self.chrome_driver_path)
            driver = webdriver.Chrome(service=service, options=options)
            
            # Test connection quickly
            driver.set_page_load_timeout(60)
            driver.get("https://www.google.com")
            return driver
                
        except Exception as e:
            self.log_message(f"⚠️ Browser setup error: {str(e)}")
            return None
    
    def extract_data(self, driver, url):
        max_retries = 3
        for attempt in range(max_retries):
            try:
                driver.set_page_load_timeout(60)
                driver.get(url)
                
                data = {}
                
                # Fast element extraction without waits
                def quick_extract(xpath):
                    try:
                        return driver.find_element(By.XPATH, xpath).text.strip()
                    except:
                        return "N/A"
                
                data["Job Title"] = quick_extract("//h1")
                data["Category"] = quick_extract('//h4[text()="دسته‌بندی شغلی"]/following-sibling::div/span')
                data["Location"] = quick_extract('//h4[text()="موقعیت مکانی"]/following-sibling::div/span')
                data["Cooperation Type"] = quick_extract('//h4[text()="نوع همکاری"]/following-sibling::div/span')
                data["Work Experience"] = quick_extract('//h4[text()="حداقل سابقه کار"]/following-sibling::div/span')
                data["Salary"] = quick_extract('//h4[text()="حقوق"]/following-sibling::div/span')
                data["Languages"] = quick_extract('//h4[text()="زبان‌های مورد نیاز"]/following-sibling::div/span')
                
                try:
                    data["Skills"] = ", ".join([skill.text.strip() for skill in driver.find_elements(By.XPATH, '//h4[text()="مهارت‌های مورد نیاز"]/following-sibling::div/span')])
                except:
                    data["Skills"] = "N/A"
                    
                data["Gender"] = quick_extract('//h4[text()="جنسیت"]/following-sibling::div/span')
                data["Military Status"] = quick_extract('//h4[text()="وضعیت نظام وظیفه"]/following-sibling::div/span')
                data["Education Level"] = quick_extract('//h4[text="حداقل مدرک تحصیلی"]/following-sibling::div/span')

                try:
                    data["Job Description"] = driver.find_element(By.CSS_SELECTOR, ".o-box__text.s-jobDesc").text.strip()
                except:
                    data["Job Description"] = "N/A"

                try:
                    data["Company Introduction"] = driver.find_element(By.CSS_SELECTOR, ".o-box__text:not(.s-jobDesc)").text.strip()
                except:
                    data["Company Introduction"] = "N/A"

                return data

            except Exception as e:
                self.log_message(f"⚠️ Attempt {attempt + 1} failed for {url}: {str(e)}")
                if attempt == max_retries - 1:
                    self.log_message(f"❌ Failed to extract data from {url} after {max_retries} attempts.")
                    return None
    
    def save_backup(self):
        try:
            backup_dir = "backups"
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(backup_dir, f"backup_{timestamp}.xlsx")
            shutil.copy(self.output_file, backup_file)
        except Exception as e:
            self.log_message(f"⚠️ Backup error: {str(e)}")

    def run_scraping(self):
        try:
            wb_input = openpyxl.load_workbook(self.input_file)
            ws_input = wb_input.active
            if ws_input is None:
                self.log_message("⚠️ Input worksheet is empty or invalid!")
                return
            
            links = [row[-1].value for row in ws_input.iter_rows(min_row=2, max_row=ws_input.max_row) if row[-1].value]
            total_links = len(links)
            
            if total_links == 0:
                self.log_message("⚠️ No links found in input file!")
                return
            
            last_processed = 0
            if self.resume_var.get() == 1:
                last_processed = self.load_status()
                if last_processed > 0:
                    self.log_message(f"Resuming from previous position: record {last_processed}")
            
            self.update_progress(last_processed, total_links)
            
            if os.path.exists(self.output_file):
                wb_output = openpyxl.load_workbook(self.output_file)
                ws_output = wb_output.active
                if ws_output is None:
                    ws_output = wb_output.create_sheet(title="Extracted Data")
            else:
                wb_output = openpyxl.Workbook()
                ws_output = wb_output.active
                if ws_output:
                    ws_output.title = "Extracted Data"
                    headers = [
                        "Job Title", "Category", "Location", "Cooperation Type", "Work Experience",
                        "Salary", "Languages", "Skills", "Gender", "Military Status",
                        "Education Level", "Job Description", "Company Introduction", "URL"
                    ]
                    ws_output.append(headers)
                else:
                    self.log_message("⚠️ Failed to create worksheet in output file!")
                    return
            
            self.log_message("Initializing browser...")
            driver = self.setup_driver()
            
            if not driver:
                self.log_message("⚠️ Browser initialization failed!")
                return
            
            for i in range(last_processed, total_links):
                if not self.is_running:
                    break
                
                # اضافه کردن تاخیر تصادفی بین درخواست‌ها
                if i > last_processed:
                    delay = random.uniform(1, 3)  # تاخیر تصادفی بین 1 تا 3 ثانیه
                    self.log_message(f"⏳ Waiting for {delay:.1f} seconds before next request...")
                    time.sleep(delay)
                
                link = links[i]
                self.update_progress(i + 1, total_links)
                self.log_message(f"Processing link {i+1}: {link}")
                
                data = self.extract_data(driver, link)
                if data and ws_output:
                    ws_output.append([
                        data["Job Title"], data["Category"], data["Location"], data["Cooperation Type"],
                        data["Work Experience"], data["Salary"], data["Languages"], data["Skills"],
                        data["Gender"], data["Military Status"], data["Education Level"],
                        data["Job Description"], data["Company Introduction"], link
                    ])
                else:
                    ws_output.append([
                        "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A",
                        "Failed to extract data", "Failed to extract data", link
                    ])
                    self.log_message(f"⚠️ Data extraction failed for link {i+1}: {link}")
                
                wb_output.save(self.output_file)
                self.save_status(i + 1)
                
                if (i + 1) % 5 == 0:
                    self.save_backup()
            
            if driver:
                driver.quit()
            
            if self.is_running:
                self.log_message("✅ Extraction completed successfully!")
                if os.path.exists(self.status_file):
                    os.remove(self.status_file)
                messagebox.showinfo("Success", "Data extraction completed!")
            else:
                self.log_message("⏹️ Extraction stopped!")
                messagebox.showinfo("Stopped", "Extraction process stopped!")
            
        except Exception as e:
            self.log_message(f"⚠️ Runtime error: {str(e)}")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        finally:
            self.is_running = False
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
            self.status_label.config(text="Ready to start again")

if __name__ == "__main__":
    root = tk.Tk()
    app = JobinjaScraperApp(root)
    root.mainloop()